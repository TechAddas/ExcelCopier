import pandas as pd
import mywritter as mw
from openpyxl import Workbook,load_workbook

outputfilename="Output_Sheet_original.xlsx"

def createDfFromExcel(input_file,filter_file,filterColumnName,inputsheetid,inPutColumnSheetName=None):
    input = pd.read_excel(input_file.get(),sheet_name=inPutColumnSheetName)
    site_list = getFilterColumnValueList(filter_file)
    data = pd.DataFrame(input[input[filterColumnName].isin(site_list)], \
        columns=  getoutPutColumnSName(inputsheetid)).reset_index().sort_values(
            by=filterColumnName,
            ascending=False,
            kind="mergesort")

    if inputsheetid ==1:
        mw.makeoutputsheetblank(getoutPutSheetName(inputsheetid),1000,17)
        createOutputSheetforInput1(data,getoutPutSheetName(inputsheetid))

    if inputsheetid ==2:
        mw.makeoutputsheetblank(getoutPutSheetName(2),1000,17)
        createOutputSheetforInput2(data,getoutPutSheetName(inputsheetid),(int(inputsheetid)-1))

    if inputsheetid ==3:
        createOutputSheetforInput2(data,getoutPutSheetName(2),(int(inputsheetid)-1))
        mw.makeoutputsheetblank(getoutPutSheetName(inputsheetid),100,8)
        createOutputSheetforInput2(data,getoutPutSheetName(3),inputsheetid)

def getoutPutColumnSName(inputid):
    if inputid == 1:
        return ["Site ID","Input A","Input B","Input C","Input D","Input E","Input F","Input G","Input H","Input I"]
    if inputid == 2:
        return ["Site ID","OS Interface IP","OS next hop IP","OS VLAN"]
    if inputid == 3:
        return ["Site ID","Core Interface IP","Core next hop IP","Core VLAN"]

def getoutPutSheetName(inputid):
    if inputid == 1:
        return "From_Input sheet1"
    if inputid == 2:
        return "From Input Sheet2"
    if inputid == 3:
        return "No input required (Fix sheet)"

def getFilterColumnValueList(filter_file_path):
    site_list = pd.read_excel (filter_file_path.get())
    return site_list["Site_List"].values.tolist()

def createOutputSheetforInput1(data,outputsheetname):
    myworkbook=load_workbook(outputfilename)
    worksheet= myworkbook[outputsheetname]
    rowindexvalue=0
    for rowindex, row in data.iterrows():
        rowindexvalue+=1

        worksheet.cell(row=int(rowindexvalue)+4,column=1).value=row["Site ID"]
        worksheet.cell(row=int(rowindexvalue)+4,column=9).value=row["Input A"]
        worksheet.cell(row=int(rowindexvalue)+4,column=10).value=row["Input B"]
        worksheet.cell(row=int(rowindexvalue)+4,column=11).value=row["Input C"]
        worksheet.cell(row=int(rowindexvalue)+4,column=12).value=row["Input D"]
        worksheet.cell(row=int(rowindexvalue)+4,column=13).value=row["Input E"]
        worksheet.cell(row=int(rowindexvalue)+4,column=14).value=row["Input F"]
        worksheet.cell(row=int(rowindexvalue)+4,column=15).value=row["Input G"]
        worksheet.cell(row=int(rowindexvalue)+4,column=16).value=row["Input H"]
        worksheet.cell(row=int(rowindexvalue)+4,column=17).value=row["Input I"]

        worksheet.cell(row=int(rowindexvalue)+4,column=2).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","Bose_band")'
        worksheet.cell(row=int(rowindexvalue)+4,column=3).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","20.Q3")'
        worksheet.cell(row=int(rowindexvalue)+4,column=7).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","template name")'


    rowindexvalue=0
    myworkbook.save(outputfilename)
    myworkbook.close()

def createOutputSheetforInput2(data,outputsheetname,sheetId):
    myworkbook=load_workbook(outputfilename)
    worksheet= myworkbook[outputsheetname]
    rowindexvalue=0
    for rowindex, row in data.iterrows():
        rowindexvalue+=1
        if sheetId ==1:

            worksheet.cell(row=int(rowindexvalue)+4,column=1).value=row["Site ID"]
            worksheet.cell(row=int(rowindexvalue)+4,column=11).value=row["OS Interface IP"]
            worksheet.cell(row=int(rowindexvalue)+4,column=12).value=row["OS next hop IP"]
            worksheet.cell(row=int(rowindexvalue)+4,column=15).value=row["OS VLAN"]
    
            worksheet.cell(row=int(rowindexvalue)+4,column=2).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","Bose_band")'
            worksheet.cell(row=int(rowindexvalue)+4,column=3).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","20.Q3")'
            worksheet.cell(row=int(rowindexvalue)+4,column=7).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","template name_2")'
            worksheet.cell(row=int(rowindexvalue)+4,column=8).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","SubNetwork=ABCD,SubNetwork=RADIOWATCH,MeetContext=Mega")'
            worksheet.cell(row=int(rowindexvalue)+4,column=9).value='=A'+str(int(rowindexvalue)+int(4))
            worksheet.cell(row=int(rowindexvalue)+4,column=10).value='=A'+str(int(rowindexvalue)+int(4))
        
        elif sheetId ==2:
            worksheet.cell(row=int(rowindexvalue)+4,column=1).value=row["Site ID"]
            worksheet.cell(row=int(rowindexvalue)+4,column=13).value=row["Core Interface IP"]
            worksheet.cell(row=int(rowindexvalue)+4,column=14).value=row["Core next hop IP"]
            worksheet.cell(row=int(rowindexvalue)+4,column=16).value=row["Core VLAN"]
        else :
            worksheet.cell(row=int(rowindexvalue)+4,column=1).value=row["Site ID"]
            worksheet.cell(row=int(rowindexvalue)+4,column=2).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","Bose_band")'
            worksheet.cell(row=int(rowindexvalue)+4,column=3).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","20.Q3")'
            worksheet.cell(row=int(rowindexvalue)+4,column=7).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","template name_2")'
            worksheet.cell(row=int(rowindexvalue)+4,column=8).value='=IF(A'+str(int(rowindexvalue)+4)+'=0,"","BBBBBBBBB")'


    rowindexvalue=0
    myworkbook.save(outputfilename)
    myworkbook.close()
